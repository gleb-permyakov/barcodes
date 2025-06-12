import pandas as pd
from itertools import product

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.formula.translate import Translator
from openpyxl.utils.dataframe import dataframe_to_rows

import math

color_art = {
    "Черный": "BLK",
    "Синий": "BLU",
    "Коричневый": "BRN",
    "Зеленый": "GRN",
    "Серый": "GRY",
    "Любой": "NCA",
    "Оранжевый": "ORG",
    "Розовый": "PNK",
    "Пурпурный": "PPL",
    "Красный": "RED",
    "Золотистый": "TAN",
    "Фиолетовый": "VIO",
    "Белый": "WHT",
    "Желтый": "YEL",
    "Прозрачный": "TRN",
    "Серебристый": "SLV",
    "Хаки": "KHK"
}

def name_to_artname(name, color, mesh_size):
    artname = []
    name = name.lower()

    if 'строительная' in name: artname.append('Building')
    elif 'сигнальная ' in name: artname.append('Signal')
    elif 'штукатурная ' in name: artname.append('Plaster')
    elif 'стяжки полов' in name: artname.append('Screed')
    elif 'кладочная' in name: artname.append('Masonry')
    elif 'дорожная' in name: artname.append('Road')
    elif 'геосетка' in name: artname.append('Geonet')
    elif 'ограждения трасс' in name: artname.append('Barrier')
    elif 'защитно-улавливающая' in name: artname.append('Safety-catch')
    elif 'временный забор' in name: artname.append('Temporary-fence')
    elif 'фасадная' in name: artname.append('Facade')
    elif 'снегозадерживающая' in name: artname.append('Snow')
    elif 'от птиц' in name: artname.append('Bird')
    elif 'от кротов' in name: artname.append('Mole')
    elif 'для цветов' in name: artname.append('Flower')
    elif 'садовая' in name: artname.append('Garden')
    elif 'под забор' in name: artname.append('Fence-base')
    elif 'заборная' in name: artname.append('Fence')
    elif 'лёгкая' in name: artname.append('Light')
    elif 'универсальная' in name: artname.append('Universal')
    elif 'пластиковая' in name: artname.append('Plastic')
    elif 'для рыбохозяйств' in name: artname.append('Fishery')
    elif 'для птичников' in name: artname.append('Poultry')
    elif 'шпалерная' in name: artname.append('Trellis')
    elif 'от зайцев' in name: artname.append('Hare')
    elif 'для гороха' in name: artname.append('Pea')
    elif 'противоградовая' in name: artname.append('Hail')
    elif 'от вытаптывания' in name: artname.append('Anti-trample')
    elif 'для гольф-полей' in name: artname.append('Golf')
    elif 'для ландшафтного дизайна' in name: artname.append('Landscape')
    elif 'для растений' in name: artname.append('Plant')
    elif 'для палисадника' in name: artname.append('Flowerbed')
    elif 'защитная' in name: artname.append('Protective')
    elif 'основа для маскировочной сети' in name: artname.append('Camouflage-base')
    elif 'основа для маскировочных костюмов' in name: artname.append('Camouflage-suit')
    elif 'для производства армированных материалов' in name: artname.append('Reinforcement')
    elif 'для производства мебели и матрасов' in name: artname.append('Furniture')
    elif 'для упаковки' in name: artname.append('Packaging')
    elif 'для сушки продуктов' in name: artname.append('Drying')
    elif 'текстильная' in name: artname.append('Textile')
    elif 'для фильтров' in name: artname.append('Filter')

    if 'упрочненная' in name: artname[0] += '-Heavy'
    elif 'облегченная' in name: artname[0] += '-Light'
    elif 'антипирен' in name: artname[0] += '-Fireproof'
    elif 'антистатик' in name: artname[0] += '-Antistatic'

    artname.append(color)
    artname.append(mesh_size)

    return "_".join(artname)

def generate_articules_with_names(combinations):
    articules = []
    for comb in combinations:
        name = comb['Название']
        color = color_art[comb['Цвет']]
        length = comb['Длина рулона (м)']
        width = comb['Ширина рулона (м)']
        if width % 1 == 0: width = int(width)
        length_width = f"{width}x{length}"
        articules.append(name_to_artname(name, color, length_width))
    return articules

def load_excel_data(file_path):
    return pd.read_excel(file_path, sheet_name="Большие рулоны", header=2)

def get_unique_combinations(df):
    required_columns = ['Категория', 'Название', 'Размер ячейки (мм)', 'Цвет', 'Вес (г/м2)', 'Ширина рулона (м)', 'Длина рулона (м)']
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"В таблице отсутствует колонка: {col}")

    unique_combs = df[required_columns].drop_duplicates()
    return unique_combs.to_dict('records')

def get_unique_combs_arts(df):
    required_columns = ['Артикул']
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"В таблице отсутствует колонка: {col}")

    unique_combs = df[required_columns].drop_duplicates()
    return unique_combs.to_dict('records')

def generate_names(combinations, result_sheet, book, output_file):
    names = []
    articules = []
    cat_5 = [] # тут мы храним категории сеток, которые идут в максимальной ширине 5м
    for comb in combinations:
        cat = comb['Категория']
        width = comb['Ширина рулона (м)']
        if (width == 5): 
            cat_5.append(cat)
    row_index = 4
    for comb in combinations:
        cat = comb['Категория']
        mesh_size = comb['Размер ячейки (мм)']
        color = comb['Цвет']
        weight = comb['Вес (г/м2)']
        width = comb['Ширина рулона (м)']
        length = comb['Длина рулона (м)']


        # записываем вариант нарезки с текущей шириной при максимально возможной 4м
        config_cut = []
        if not cat in cat_5:
            config_cut = write_config_cut(4, width)
        else: 
            config_cut = write_config_cut(5, width)

        for i in range(len(config_cut)): # убираем лишние нули в записи
            if config_cut[i] % 1 == 0: config_cut[i] = int(config_cut[i])
            config_cut[i] = str(config_cut[i])

        for i in range(len(config_cut)): # убираем лишние нули в записи, точку меняем на запятую
            if '.' in config_cut[i]: config_cut[i] = f"{float(config_cut[i]):.1f}".replace(".", ",")

        if (len(config_cut) > 5): s_config_cut = f"{config_cut[0]}x{len(config_cut)}" # чтобы писалось 0,5x8 например
        else: s_config_cut = "+".join(map(str, config_cut)) # собираем все в одну строчку через + 

        mesh_size = str(mesh_size).replace(" ", "") # затираем нули в размере ячейки

        # убираем в записи незначащие нули
        if weight % 1 == 0: weight = int(weight)
        # if width % 1 == 0: width = int(width)
        if length % 1 == 0: length = int(length)

        # точку меняем на запятую
        # width = str(width)
        length = str(length)
        # if '.' in width: width = f"{float(width):.1f}".replace(".", ",")
        if '.' in length: length = f"{float(length):.1f}".replace(".", ",")

        if (cat in cat_5): const_width = "5" # ширина в артикуле 5
        else: const_width = "4" # ширина в артикуле 4

        # добавляем имя
        name = (f"Сетка полимерная, {mesh_size}, {color}, {weight} г/м2, {const_width}x{length} ({s_config_cut})")

        names.append(name)
        # добавляем артикул
        if (cat in cat_5): art_width = "5" # ширина в артикуле 5
        else: art_width = "4" # ширина в артикуле 4
        art = generate_arts(color, mesh_size, weight, art_width, length, config_cut)
        articules.append(art)

        # 2 + 2 -> 1 + 1 + 2
        if (len(config_cut) == 2 and config_cut[0] == '2' and config_cut[1] == '2'): 

            row_to_copy = result_sheet[row_index]

            # Меняем нужные значения в новой строке
            result_sheet.cell(row=row_index + 1, column=2).value = generate_arts(color, mesh_size, weight, art_width, length, [1,1,2])# артикул
            result_sheet.cell(row=row_index + 1, column=4).value = f"Сетка полимерная, {mesh_size}, {color}, {weight} г/м2, {const_width}x{length} (1+1+2)"  # имя

            articules.append(generate_arts(color, mesh_size, weight, art_width, length, [1,1,2]))
            names.append(f"Сетка полимерная, {mesh_size}, {color}, {weight} г/м2, {const_width}x{length} (1+1+2)")

            # Вставляем новую строку
            result_sheet.insert_rows(row_index + 1)

            # Копируем данные и стили из исходной строки в новую
            for col_idx, source_cell in enumerate(row_to_copy, start=1):
                target_cell = result_sheet.cell(row=row_index + 1, column=col_idx)
                
                # Копируем значение
                target_cell.value = source_cell.value
                
                # Копируем стиль (если он есть)
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
            row_index += 1

        row_index += 1
    print(len(names), len(articules))

    return [names, articules]


def generate_arts(color, mesh_size, weight, art_width, length, config_cut):
    if len(config_cut) > 5:  s_config_cut = f"{ str(config_cut[0]) }".replace(",", "")
    else: 
        s_config_cut = ""
        for i in config_cut: s_config_cut += str(i).replace(",", "")
    
    return f"{color_art[color]}_{mesh_size}_{weight}_{art_width}_{length}_{s_config_cut}"


def generate_names_2(combinations):
    names = []
    articules = []
    cat_5 = [] # тут мы храним категории сеток, которые идут в максимальной ширине 5м
    for comb in combinations:
        cat = comb['Категория']
        width = comb['Ширина рулона (м)']
        if (width == 5): 
            cat_5.append(cat)
    for comb in combinations:
        cat = comb['Категория']
        mesh_size = comb['Размер ячейки (мм)']
        color = comb['Цвет']
        weight = comb['Вес (г/м2)']
        width = comb['Ширина рулона (м)']
        length = comb['Длина рулона (м)']

        # записываем вариант нарезки с текущей шириной при максимально возможной 4м
        config_cut = []
        if not cat in cat_5:
            config_cut = write_config_cut(4, width)
        else: 
            config_cut = write_config_cut(5, width)

        for i in range(len(config_cut)): # убираем лишние нули в записи
            if config_cut[i] % 1 == 0: config_cut[i] = int(config_cut[i])
            config_cut[i] = str(config_cut[i])

        for i in range(len(config_cut)): # убираем лишние нули в записи, точку меняем на запятую
            if '.' in config_cut[i]: config_cut[i] = f"{float(config_cut[i]):.1f}".replace(".", ",")

        if (len(config_cut) > 5): s_config_cut = f"{config_cut[0]}x{len(config_cut)}" # чтобы писалось 0,5x8 например
        else: s_config_cut = "+".join(map(str, config_cut)) # собираем все в одну строчку через +

        mesh_size = str(mesh_size).replace(" ", "") # затираем нули в размере ячейки

        # убираем в записи незначащие нули
        if weight % 1 == 0: weight = int(weight)
        if width % 1 == 0: width = int(width)
        if length % 1 == 0: length = int(length)

        # точку меняем на запятую
        width = str(width)
        length = str(length)
        if '.' in width: width = f"{float(width):.1f}".replace(".", ",")
        if '.' in length: length = f"{float(length):.1f}".replace(".", ",")

        # добавляем имя
        name = (f"Сетка полимерная, {mesh_size}, {color}, {weight} г/м2, {width}x{length}м")
        names.append(name)
        # добавляем артикул
        if (cat in cat_5): art_width = "5" # ширина в артикуле 5
        else: art_width = "4" # ширина в артикуле 4
        art = generate_arts_2(color, mesh_size, weight, width, length)
        articules.append(art)
    
    return [names, articules]


def generate_arts_2(color, mesh_size, weight, width, length):
    return (f"{color_art[color]}_{mesh_size}_{weight}_{width}_{length}").replace(",", "")


def write_config_cut(max_width, width):
    var_widths = []
    for i in range (int(max_width // width)):
        var_widths.append(width)
    # print(max_width, width, max_width // width)
    extra = max_width % (width * (max_width // width))
    if extra != 0: var_widths.append(extra)
    return list(reversed(var_widths))


def process_excel_file(input_file, output_file):

    df = load_excel_data(input_file)

    # Загружаем существующую книгу
    book = load_workbook(input_file)

    # базовый лист
    base_sheet = book['Большие рулоны']

    # ЛИСТ СО СГЕНЕРИРОВАННЫМИ АРТИКУЛАМИ
    if 'Результаты' in book.sheetnames:
        del book['Результаты']

    result_sheet = book.copy_worksheet(book['Большие рулоны'])
    result_sheet.title = 'Результаты'

    unique_combs = get_unique_combinations(df)
    # print(len(unique_combs), unique_combs[0])

    # тут будем вставлять артикул по имени, цвету и клетке
    articules_with_names = generate_articules_with_names(unique_combs)
    insert_names_to_column(base_sheet, articules_with_names, column_name='Артикул')

    print(f"Всего строк в df: {len(df)}")   
    print(f"Уникальных комбинаций: {len(unique_combs)}")
    duplicates = df[df.duplicated(keep=False)]  # keep=False помечает все дубли
    # print(f"Дублирующиеся строки:\n{duplicates}")

    # Сохраняем изменения
    book.save(output_file)

    [generated_names, articules] = generate_names(unique_combs, result_sheet, book, output_file)
    
    # Вставляем названия в столбец "Название"
    insert_names_to_column(result_sheet, generated_names, column_name='Название')
    # Вставляем артикулы в столбец "Артикул"
    insert_names_to_column(result_sheet, articules, column_name='Артикул')
    
    # result_sheet.delete_cols(find_col_index(result_sheet, "Категория"))
    # Можно удалить, но все сломается, поэтому сузим столбец
    # Получаем буквенное обозначение столбца и меняем ширину
    column_letter = get_column_letter(2)
    result_sheet.column_dimensions[column_letter].width = 0
    column_letter = get_column_letter(1)
    result_sheet.column_dimensions[column_letter].width = 42
    column_letter = get_column_letter(3)
    result_sheet.column_dimensions[column_letter].width = 45

    # Сохраняем изменения
    book.save(output_file)

    # # убираем строки с дублирующимся артикулом - последнее действие!!
    # unique_table = get_unique_combs_arts(df)
    # # Очистка листа (если нужно оставить шапку, удалите эту строку)
    # result_sheet.delete_rows(3, base_sheet.max_row)

    # # Записываем данные
    # for i, record in enumerate(unique_table, start=3):
    #     base_sheet.cell(row=i, column=1, value=record['Артикул'])

    # # Сохранение изменений
    # book.save(input_file)




    # ЛИСТ С НАРЕЗАННЫМИ РУЛОНАМИ
    if 'Нарезанные рулоны' in book.sheetnames:
        del book['Нарезанные рулоны']

    result_sheet = book.copy_worksheet(book['Большие рулоны'])
    result_sheet.title = 'Нарезанные рулоны'

    [generated_names, articules] = generate_names_2(unique_combs)

    # Сохраняем изменения
    book.save(output_file)
    # Вставляем названия в столбец "Название"
    insert_names_to_column(result_sheet, generated_names, column_name='Название')
    # Вставляем артикулы в столбец "Артикул"
    insert_names_to_column(result_sheet, articules, column_name='Артикул')
    
    # result_sheet.delete_cols(find_col_index(result_sheet, "Категория"))
    # Можно удалить, но все сломается, поэтому сузим столбец
    # Получаем буквенное обозначение столбца и меняем ширину
    column_letter = get_column_letter(2)
    result_sheet.column_dimensions[column_letter].width = 0
    column_letter = get_column_letter(1)
    result_sheet.column_dimensions[column_letter].width = 42
    column_letter = get_column_letter(3)
    result_sheet.column_dimensions[column_letter].width = 45

    # Сохраняем изменения
    book.save(output_file)

    print(f"Все сохранено в {output_file}")


def insert_names_to_column(sheet, names, column_name='Название'):
    """Вставляет названия в указанный столбец"""
    # Находим индекс столбца
    col_idx = None
    for cell in sheet[3]:  # Ищем в заголовках
        if cell.value and str(cell.value).lower() == column_name.lower():
            col_idx = cell.column
            break
    
    if col_idx is None:
        raise ValueError(f"Столбец '{column_name}' не найден")
    
    # Вставляем данные (начиная со второй строки)
    for i, name in enumerate(names, start=4):
        sheet.cell(row=i, column=col_idx, value=name)


def find_col_index(sheet, col_name):
    # Находим индекс столбца
    col_id = None
    for cell in sheet[2]:  # Ищем в заголовках
        if cell.value and str(cell.value).lower() == col_name.lower():
            col_id = cell.column
            break
    
    if col_id is None:
        raise ValueError(f"Столбец '{col_name}' не найден")
    
    return col_id


if __name__ == "__main__":
    input_excel = "data.xlsx"
    output_excel = "result.xlsx"
    
    process_excel_file(input_excel, input_excel)